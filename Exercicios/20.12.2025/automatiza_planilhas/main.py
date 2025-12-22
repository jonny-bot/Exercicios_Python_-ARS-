from openpyxl import load_workbook, Workbook

# Ler Dados de Qualquer Planilha
planilha_vendas = load_workbook('vendas_de_lanches.xlsx')
pagina_vendas = planilha_vendas['Sheet']

for linha in pagina_vendas.iter_rows(values_only=True):
    print(linha)

# Automatizar Entrada de Dados em Planilhas
# Inserir Dados de Qualquer Fonte(Word, Banco de Dados, Outros Sistemas)
planilhas_contas = Workbook()
pagina1 = planilhas_contas.active

with open('anotações.txt','r',encoding='utf-8') as arquivo:
    for linha in arquivo:
        pagina1.append(linha.split(','))

planilhas_contas.save('contas_pagar.xlsx')
